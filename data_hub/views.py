from django.shortcuts import render
from django.db import connection
import openpyxl
import itertools


# Create your views here.
def createtab(request):
    if request.method == 'POST':
        # if request.POST.get('x'):
        #     cursor=connection.cursor()
        #     cursor.execute(request.POST.get('x'))
        #     return render(request, 'index.html')
        # else:

        # https: // pythoncircle.com / post / 591 / how - to - upload - and -process - the - excel - file - in -django /
        excel_file = request.FILES["excel_file"]
        table_name = request.POST.get('table_name')

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]
        # print(worksheet)

        excel_data_1 = list()
        excel_data_2 = list()
        excel_data_3 = list()

        # iterating over the rows and
        # getting value from each cell in row

        # to get the type of fields.
        for row in worksheet.iter_rows(max_row=1):
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data_1.append(row_data)
        excel_data_1 = excel_data_1[0]
        # excel_data_1 = tuple(excel_data_1)
        # excel_data_1 = ' '.join([str(elem) for elem in excel_data_1])
        # print('XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX')
        # print(excel_data_1)
        # print('XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX')

        # to get the field's name.
        for row in worksheet.iter_rows(min_row=2, max_row=2):
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data_2.append(row_data)
        excel_data_2 = tuple(excel_data_2)
        excel_data_2 = excel_data_2[0]
        # print('XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX')
        # print(excel_data_2)
        # print('XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX')

        # to get row values.
        for row in worksheet.iter_rows(min_row=3):
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data_3.append(row_data)
        # excel_data_3 = excel_data_3[0]
        # print('XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX')
        # print(excel_data_3)
        # print('XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX')

        # converting data into postgres commands:-
        create_table_name = "CREATE TABLE " + str(table_name)
        # print('XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX')
        # print(create_table_name)
        # print('XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX')

        # # Python3 program to convert a
        # # list into a tuple
        # def convert(list):
        #     return tuple(list)

        _create_columns_dictionary = dict(zip(tuple(excel_data_2), tuple(excel_data_1)))
        # _create_columns_dictionary = dict(map(lambda i, j: (i, j), excel_data_1, excel_data_2))
        # _create_columns_dictionary = {}
        # for i in excel_data_1:
        #     for j in excel_data_2:
        #         _create_columns_dictionary[i] = j
        # # for (i, j) in zip(excel_data_1, excel_data_2):
        # #     print(i, j)
        # print('XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX')
        # print(_create_columns_dictionary)
        # print('XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX')

        # create_columns = []
        # for i, j in _create_columns_dictionary.items():
        #     create_columns.append(
        #         str(j) + " " +str(i) + "()"
        #     )
        # print('XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX')
        # print(create_columns)
        # print('XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX')

        create_columns = ''
        for _column, _datatype in _create_columns_dictionary.items():
            create_columns += "`" + _column + "`" + " " + _datatype + " ,"
        create_columns = create_columns[0:-2]

        # print('XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX')
        # print(create_columns)
        # print('XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX')

        create_table_command = create_table_name + "(" + create_columns + ");"

        # print('XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX')
        # print(create_table_command)
        # print('XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX')

        # creating commands to populate data into table
        insert_into_command = "INSERT INTO " + table_name + " "
        # print('XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX')
        # print(insert_into_command)
        # print('XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX')

        columns = ''
        for _ in _create_columns_dictionary.keys():
            columns += _ + ", "
        columns = columns[0:-2]
        # print('XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX')
        # print(columns)
        # print('XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX')

        _val = ''
        _k = ''
        for i in excel_data_3:
            for j in i:
                _k += "'" + j + "',"
            # _val += str("(" + _k + ")",)
            _k = _k[0:-1]
            _val += "(" + _k + "), "
            _k = ''
        _val = _val[0:-2]
        rows = _val + ";"

        # print('XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX')
        # print(rows)
        # print('XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX')

        # creating command to populate the table
        insert_into_table_command = insert_into_command + '(' + columns + ') ' + 'VALUES ' + rows
        # print('XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX')
        # print(insert_into_table_command)
        # print('XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX')

        # # creating table


        # # to create a new table
        # cursor.execute(create_table_command)
        #
        # # to populate data into a created table
        # cursor.execute(insert_into_table_command)

        # print('XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX')
        # print(create_table_command)
        # print('XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX')
        #
        # print('XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX')
        # print(insert_into_table_command)
        # print('XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX')

        return render(request, 'index.html')
    else:
        return render(request, 'index.html')