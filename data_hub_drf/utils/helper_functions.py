import openpyxl
import json

from openpyxl.utils.exceptions import InvalidFileException
from rest_framework import status
from rest_framework.response import Response

from data_hub_drf.utils.custom_exceptions import InvalidPayload
from data_hub_drf.forms import UploadForm
from data_hub_drf.utils.error_handler import error_message


def is_excel_file(file):
    """to check if uploaded file is excel or not and if it is then load into memory."""
    try:
        wb = openpyxl.load_workbook(file)
        return True, wb
    except:
        return False


def table_name_worksheet_verifier(request=None):
    """To get the name of the excel and check if "Sheet1" is present in the excel file or not and if there are more
    then one sheet then yield error. """

    form = UploadForm(request.POST, request.FILES)

    if form.is_valid():
        excel_file = request.FILES["file"]
        try:
            # when is_excel_file = True, is also returning wb
            file_type_excel, wb = is_excel_file(excel_file)
        except:
            # when is_excel_file = False, is only returning False
            file_type_excel = is_excel_file(excel_file)

        if file_type_excel:
            file_name = str(excel_file)

            table_name, _ = file_name.rsplit('.', 1)
            if len(wb.worksheets) > 1:
                raise InvalidPayload(detail='More than one Sheet provided.')
            if len(wb.worksheets) < 1:
                raise InvalidPayload(detail='No Sheet provided.')
            if wb.worksheets[0].title != "Sheet1":
                raise InvalidPayload(detail='Provide worksheet name as Sheet1.')
            worksheet = wb["Sheet1"]
            return worksheet, table_name
        else:
            raise InvalidPayload(detail='Uploaded file is not excel.')
    return form


def data_extractor(worksheet=None, data_types=None):
    """
    To extract all the data from the workbook.

    if "data_types" are provided then, will be populating the list "row_data_1" by "data_types".
    """

    # row_data_1: field types, row_data_2: field names, row_data_3: row_data_3
    row_data_1 = []
    row_data_2 = []
    row_data_3 = []

    # to get the field names
    for row in worksheet.iter_rows(max_row=1):
        for cell in row:
            row_data_2.append(str(cell.value))

    if data_types == None:
        # to get the field types
        for row in worksheet.iter_rows(min_row=2, max_row=2):
            for cell in row:
                row_data_1.append(str(cell.value))
    else:
        # data_types = json.loads(data_types)
        try:
            data_types = json.loads(data_types)
        except json.decoder.JSONDecodeError as e:
            _error = f"An error occurred: {e}"
            return _error

        error_fields = []
        # to get the field types
        # creating the data_type list (row_data_1) in ordered way w.r.t. to the field name list (row_data_2)
        for i in row_data_2:
            try:
                row_data_1.append(data_types[i])
            except:
                error_fields.append(i)
                _error = f"Data type of fields {error_fields} are not defined"
        # to check if _error is declared or not (exists or not)
        if "_error" in locals():
            return _error

        # importing state_handlers to perform appropriate action for populating data into created table
        from data_hub_drf.utils.state_handlers import (
            UPLOAD_EXCEL_METHOD_1,
            UPLOAD_EXCEL_METHOD_2,
            UPLOAD_EXCEL_METHOD_1_DATA_START_ROW_3,
            UPLOAD_EXCEL_METHOD_2_DATA_START_ROW_2
        )

        if UPLOAD_EXCEL_METHOD_1:
            min_row = UPLOAD_EXCEL_METHOD_1_DATA_START_ROW_3
        else:
            min_row = UPLOAD_EXCEL_METHOD_2_DATA_START_ROW_2

        for row in worksheet.iter_rows(min_row=min_row):
            row_data = []
            for cell in row:
                row_data.append(str(cell.value))
            row_data_3.append(row_data)

    return row_data_1, row_data_2, row_data_3